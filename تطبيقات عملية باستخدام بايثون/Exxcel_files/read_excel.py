import openpyxl
from pathlib import Path
excel=openpyxl.load_workbook(Path.home()/Path('Desktop','emplyees.xlsx'))
print(excel.sheetnames)

sheet_1=excel['Sheet1']
print(sheet_1.title)


print(excel.active.title)


print(sheet_1['A2'].value)
print(sheet_1['B2'].value)
print(sheet_1['c2'].value)
print(sheet_1['c2'].row)
print(sheet_1['c2'].column)
print(sheet_1['c2'].coordinate)
print(sheet_1.cell(row=5,column=2).value)

for i in range(2,10):
    print(f'{i}',sheet_1.cell(row=i,column=1).value)

print('-'*50)    

x=0
totale=0
for i in range(1,sheet_1.max_row):
    x+=1
    if x>=2:
        if sheet_1.cell(row=i,column=1).value!=None:
            names=f'{i-1}  {sheet_1.cell(row=i,column=1).value}'
            ages=f',{sheet_1.cell(row=i,column=2).value} ages' 
            salary=f'{i-1}{sheet_1.cell(row=i,column=3).value}$'
            num=20-len(names)
            space_name=' '*num
            num=20-len(ages)
            space_ages=' '*num
            all=f'{names}{space_name}{ages}{space_ages}{salary}'
            totale+=int(sheet_1.cell(row=i,column=3).value)

        else:
            break

    else:                    
        names=f'',sheet_1.cell(row=i,column=1).value
        ages=f'',sheet_1.cell(row=i,column=2).value
        salary=f'',sheet_1.cell(row=i,column=3).value,'$'



    print(all)
    
print(f'the totale of salary of the employees is {totale} $')


print('-'*50)

