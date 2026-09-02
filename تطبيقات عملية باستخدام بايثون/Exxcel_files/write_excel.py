import openpyxl
from pathlib import Path
#creat the sheet 
excel_file=openpyxl.Workbook()


print(excel_file.sheetnames)

excel=excel_file.active
excel.title='first_sheet'
print(excel.title)
#save the sheet 


excel_file.create_sheet()
excel_file.create_sheet()
excel_file.create_sheet()
print(excel_file.sheetnames) 

excel_file.create_sheet(index=0,title='one_sheet')
excel_file.create_sheet(index=1,title='two_sheet')
del excel_file['Sheet2']
print(excel_file.sheetnames) 
print(excel_file.sheetnames) 
print(excel_file.sheetnames) 

#write_in_a_sheet
sheet_1=excel_file['Sheet1']
names_list=['yousef','ahmad','salman','jaber']
for i in range(1,len(names_list)+1):
    sheet_1.cell(row=i,column=3).value=names_list[i-1]
    print(sheet_1.cell(row=i,column=3).value)













excel_file.save(filename=Path.home()/Path('Desktop','write_excel.xlsx'))

