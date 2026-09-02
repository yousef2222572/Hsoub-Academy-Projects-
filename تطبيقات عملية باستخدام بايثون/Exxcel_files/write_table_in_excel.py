import openpyxl
from openpyxl.styles import Font
import sys
from pathlib import Path




if len(sys.argv)==2:
    excel=openpyxl.Workbook()
    excel.create_sheet()
    sheet=excel.sheetnames
    excel.worksheets[0].title='Times'
    sheet1=excel.worksheets[0]

    for i in range(1,int(sys.argv[1])+1):
        sheet1.cell(row=i+1,column=1).font=Font(size=20)
        sheet1.cell(row=1,column=i+1).font=Font(size=20)
        sheet1.cell(row=i+1,column=1).value=i
        sheet1.cell(row=1,column=i+1).value=i
        for x in range(1,int(sys.argv[1])+1):
            sheet1.cell(row=x+1,column=i+1).value=i*x
name=Path.home()/Path('Desktop',f'TIMES__{sys.argv[1]}__.xlsx')
print(f'your file at ({str(name)})')
excel.save(filename=Path.home()/Path('Desktop',f'TIMES{sys.argv[1]}.xlsx'))


