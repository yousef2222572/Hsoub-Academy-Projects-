import requests
import selenium
from pathlib import Path
import csv
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl




browser = webdriver.Chrome()

browser.get('https://google.com')






browser.get("https://en.wikipedia.org/wiki/List_of_languages_by_number_of_native_speakers",)
try:

    elem=browser.find_elements(By.TAG_NAME,'table')

    elem=elem[1]
    

    body=elem.find_elements(By.TAG_NAME,'td')
    headers=elem.find_elements(By.TAG_NAME,'th')
    headers_list=[]    
    data_dict={} 
    header_length=0
    header_count=0
    header_list=[]
    b_row_list=[]
    for row in headers :
        header_list.append(row.text)
        header_length+=1
        print(row.text)
        headers_list.append(row.text)
    body_lenght=len(body)//2
    print('body lenght',body_lenght)
    cells_counter=0
    for b_row in body :
        cells_counter+=1
        
        if cells_counter==body_lenght:
            data_dict[header_list[header_count]]=b_row_list
            header_count+=1
            cells_counter=0
            b_row_list=[]
            
            
        b_row_list.append(b_row.text)
    print(data_dict)
    excel=openpyxl.Workbook()

    sheet=excel['Sheet']



    x=0
    for key,itme in data_dict.items():
        row_count=1
        
        x+=1
        sheet.cell(row=1, column=x).value = key
        
        for one_item in itme:
            row_count+=1
            sheet.cell(row=row_count, column=x).value = one_item

    excel.save('scraping_xlsx.xlsx')
            

    

    
except  :
    print('error not found ')


